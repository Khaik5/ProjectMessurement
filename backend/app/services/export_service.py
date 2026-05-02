from __future__ import annotations

import io
import json
import math
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.database import fetch_all, fetch_one
from app.repositories import dataset_repository, model_repository, report_repository

NAVY = "0F172A"
INK = "111827"
MUTED = "64748B"
WHITE = "FFFFFF"
LINE = "CBD5E1"
SOFT_BG = "F8FAFC"
PRIMARY = "2563EB"

RISK_ORDER = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
RISK_FILLS = {"LOW": "DCFCE7", "MEDIUM": "FEF3C7", "HIGH": "FFEDD5", "CRITICAL": "FEE2E2"}
RISK_TEXT = {"LOW": "166534", "MEDIUM": "854D0E", "HIGH": "9A3412", "CRITICAL": "991B1B"}
RISK_HEAT = {"LOW": "22C55E", "MEDIUM": "FACC15", "HIGH": "F97316", "CRITICAL": "EF4444"}
PDF_RISK_FILL = {
    "LOW": colors.HexColor("#DCFCE7"),
    "MEDIUM": colors.HexColor("#FEF3C7"),
    "HIGH": colors.HexColor("#FFEDD5"),
    "CRITICAL": colors.HexColor("#FEE2E2"),
}
PDF_RISK_TEXT = {
    "LOW": colors.HexColor("#166534"),
    "MEDIUM": colors.HexColor("#854D0E"),
    "HIGH": colors.HexColor("#9A3412"),
    "CRITICAL": colors.HexColor("#991B1B"),
}


@dataclass(frozen=True)
class ExportOptions:
    include_full_modules: bool = True
    include_heatmap: bool = True
    include_charts: bool = True
    top_n: int = 20


def _safe_float(value: Any, default: float | None = 0.0) -> float | None:
    if value is None:
        return default
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(numeric) or math.isinf(numeric):
        return default
    return numeric


def _safe_int(value: Any, default: int = 0) -> int:
    numeric = _safe_float(value, None)
    if numeric is None:
        return default
    return int(round(numeric))


def _probability(value: Any) -> float | None:
    numeric = _safe_float(value, None)
    if numeric is None:
        return None
    if numeric > 1 and numeric <= 100:
        numeric = numeric / 100.0
    return max(0.0, min(1.0, numeric))


def _percent_ratio(value: Any) -> float | None:
    numeric = _safe_float(value, None)
    if numeric is None:
        return None
    if numeric > 1 and numeric <= 100:
        numeric = numeric / 100.0
    return max(0.0, min(1.0, numeric))


def _risk_level(probability: float | None, fallback: Any = None) -> str:
    risk = str(fallback or "").upper()
    if risk in RISK_ORDER:
        return risk
    prob = _probability(probability) or 0.0
    if prob < 0.30:
        return "LOW"
    if prob < 0.60:
        return "MEDIUM"
    if prob < 0.80:
        return "HIGH"
    return "CRITICAL"


def _fmt_dt(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value is None:
        return "N/A"
    return str(value)


def _fmt_pct(value: Any) -> str:
    prob = _probability(value)
    return "N/A" if prob is None else f"{prob * 100:.2f}%"


def _fmt_number(value: Any, decimals: int = 2) -> str:
    numeric = _safe_float(value, None)
    if numeric is None:
        return "N/A"
    return f"{numeric:.{decimals}f}"


def _safe_text(value: Any, default: str = "N/A") -> str:
    if value is None:
        return default
    text = str(value)
    return text if text.strip() else default


def _json_loads(value: Any, default: Any = None) -> Any:
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return default


def _filename_token(value: Any) -> str:
    token = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "dataset")).strip("_")
    return token[:80] or "dataset"


def export_filename(dataset_id: int, extension: str) -> str:
    dataset = dataset_repository.get_dataset(dataset_id)
    name = _filename_token((dataset or {}).get("file_name") or f"dataset_{dataset_id}")
    stamp = datetime.now().strftime("%Y-%m-%d")
    return f"DefectAI_Report_{name}_{stamp}.{extension.lstrip('.')}"


def _dataset_model(dataset_id: int) -> dict[str, Any] | None:
    model = fetch_one(
        """
        SELECT TOP 1 m.*
        FROM Predictions p
        JOIN MLModels m ON m.id = p.model_id
        WHERE p.dataset_id = ? AND ISNULL(m.is_deleted,0)=0
        ORDER BY p.created_at DESC, p.id DESC
        """,
        [dataset_id],
    )
    return model or model_repository.get_active_production_model()


def _prediction_records(dataset_id: int) -> list[dict[str, Any]]:
    rows = fetch_all(
        """
        WITH prediction_rows AS (
            SELECT
                p.*,
                ROW_NUMBER() OVER (ORDER BY p.id ASC) AS row_no
            FROM Predictions p
            WHERE p.dataset_id = ?
        ),
        metric_rows AS (
            SELECT
                mr.*,
                ROW_NUMBER() OVER (ORDER BY mr.id ASC) AS row_no
            FROM MetricRecords mr
            WHERE mr.dataset_id = ?
        )
        SELECT
            p.id AS prediction_id,
            p.dataset_id,
            p.project_id,
            p.model_id,
            p.module_name,
            p.loc AS prediction_loc,
            p.complexity AS prediction_complexity,
            p.coupling AS prediction_coupling,
            p.code_churn AS prediction_code_churn,
            p.defect_probability,
            p.prediction,
            p.prediction_label,
            p.risk_score,
            p.defect_density,
            p.size_score,
            p.complexity_score,
            p.coupling_score,
            p.churn_score,
            r.name AS risk_level,
            r.color AS risk_color,
            p.suggested_action,
            p.created_at,
            m.name AS model_name,
            m.model_type,
            m.training_profile,
            m.threshold,
            m.accuracy,
            m.precision,
            m.recall,
            m.f1_score,
            m.roc_auc,
            m.pr_auc,
            m.selection_strategy,
            m.selection_score,
            m.metrics_json,
            mr.loc AS metric_loc,
            mr.ncloc,
            mr.cloc,
            mr.complexity AS metric_complexity,
            mr.cyclomatic_complexity,
            mr.depth_of_nesting,
            mr.coupling AS metric_coupling,
            mr.cohesion,
            mr.information_flow_complexity,
            mr.code_churn AS metric_code_churn,
            mr.change_request_backlog,
            mr.pending_effort_hours,
            mr.percent_reused
        FROM prediction_rows p
        LEFT JOIN metric_rows mr
            ON mr.dataset_id = p.dataset_id
            AND mr.row_no = p.row_no
        LEFT JOIN RiskLevels r
            ON r.id = p.risk_level_id
        LEFT JOIN MLModels m
            ON m.id = p.model_id
        ORDER BY p.defect_probability DESC, p.id ASC
        """,
        [dataset_id, dataset_id],
    )
    if not rows:
        raise ValueError("Dataset has no predictions to export.")

    normalized = []
    for row in rows:
        threshold = _safe_float(row.get("threshold"), None)
        probability = _probability(row.get("defect_probability"))
        risk = _risk_level(probability, row.get("risk_level"))
        predicted = row.get("prediction")
        if predicted is None:
            predicted = int((probability or 0.0) >= (threshold if threshold is not None else 0.5))
        normalized.append(
            {
                "prediction_id": row.get("prediction_id"),
                "dataset_id": row.get("dataset_id"),
                "project_id": row.get("project_id"),
                "module_name": _safe_text(row.get("module_name"), ""),
                "defect_probability": probability,
                "prediction_label": int(bool(predicted)),
                "prediction_text": _safe_text(row.get("prediction_label"), "Defect" if predicted else "No Defect"),
                "risk_level": risk,
                "risk_score": _probability(row.get("risk_score")),
                "suggested_action": _safe_text(row.get("suggested_action"), ""),
                "model_source": "AI production model" if row.get("model_id") else "Measurement fallback",
                "model_id": row.get("model_id"),
                "model_name": _safe_text(row.get("model_name"), "Measurement fallback"),
                "model_type": _safe_text(row.get("model_type"), "fallback"),
                "training_profile": _safe_text(row.get("training_profile"), "N/A"),
                "threshold": threshold if threshold is not None else 0.5,
                "created_at": row.get("created_at"),
                "loc": _safe_float(row.get("metric_loc"), row.get("prediction_loc")),
                "ncloc": _safe_float(row.get("ncloc"), None),
                "cloc": _safe_float(row.get("cloc"), None),
                "complexity": _safe_float(row.get("metric_complexity"), row.get("prediction_complexity")),
                "cyclomatic_complexity": _safe_float(row.get("cyclomatic_complexity"), None),
                "depth_of_nesting": _safe_float(row.get("depth_of_nesting"), None),
                "coupling": _safe_float(row.get("metric_coupling"), row.get("prediction_coupling")),
                "cohesion": _percent_ratio(row.get("cohesion")),
                "information_flow_complexity": _safe_float(row.get("information_flow_complexity"), None),
                "code_churn": _safe_float(row.get("metric_code_churn"), row.get("prediction_code_churn")),
                "change_request_backlog": _safe_float(row.get("change_request_backlog"), None),
                "pending_effort_hours": _safe_float(row.get("pending_effort_hours"), None),
                "percent_reused": _percent_ratio(row.get("percent_reused")),
            }
        )
    return normalized


def _summary(dataset: dict[str, Any], rows: list[dict[str, Any]], model: dict[str, Any] | None) -> dict[str, Any]:
    probabilities = [row["defect_probability"] for row in rows if row.get("defect_probability") is not None]
    total = len(rows)
    predicted = sum(1 for row in rows if int(row.get("prediction_label") or 0) == 1)
    risk_counts = Counter(row.get("risk_level") for row in rows)
    created_values = [row.get("created_at") for row in rows if row.get("created_at")]
    analysis_date = max(created_values) if created_values else None
    threshold = _safe_float((model or {}).get("threshold"), None)
    if threshold is None and rows:
        threshold = rows[0].get("threshold")
    return {
        "project_id": dataset.get("project_id"),
        "dataset_id": dataset.get("id"),
        "dataset_name": dataset.get("file_name") or dataset.get("name") or dataset.get("id"),
        "analysis_date": analysis_date,
        "active_model": _safe_text((model or {}).get("name"), rows[0].get("model_name") if rows else "Measurement fallback"),
        "model_type": _safe_text((model or {}).get("model_type"), rows[0].get("model_type") if rows else "fallback"),
        "training_profile": _safe_text((model or {}).get("training_profile"), rows[0].get("training_profile") if rows else "N/A"),
        "threshold": threshold if threshold is not None else 0.5,
        "total_modules": total,
        "predicted_defect_modules": predicted,
        "defect_rate": (predicted / total) if total else 0.0,
        "avg_defect_probability": (sum(probabilities) / len(probabilities)) if probabilities else 0.0,
        "highest_defect_probability": max(probabilities) if probabilities else 0.0,
        "lowest_defect_probability": min(probabilities) if probabilities else 0.0,
        "low_count": risk_counts.get("LOW", 0),
        "medium_count": risk_counts.get("MEDIUM", 0),
        "high_count": risk_counts.get("HIGH", 0),
        "critical_count": risk_counts.get("CRITICAL", 0),
        "high_critical_count": risk_counts.get("HIGH", 0) + risk_counts.get("CRITICAL", 0),
        "model_source": rows[0].get("model_source") if rows else "Measurement fallback",
        "generated_at": datetime.now(),
    }


def _build_report(dataset_id: int) -> dict[str, Any]:
    dataset = dataset_repository.get_dataset(dataset_id)
    if not dataset:
        raise ValueError(f"Dataset #{dataset_id} not found.")
    rows = _prediction_records(dataset_id)
    model = _dataset_model(dataset_id)
    return {
        "dataset": dataset,
        "rows": rows,
        "model": model,
        "summary": _summary(dataset, rows, model),
        "models": _model_rows(dataset_id, model),
    }


def _model_rows(dataset_id: int, active_model: dict[str, Any] | None) -> list[dict[str, Any]]:
    models = model_repository.list_models()
    scoped = [
        model
        for model in models
        if model.get("dataset_id") in (dataset_id, None) or (active_model and model.get("id") == active_model.get("id"))
    ]
    return scoped or ([active_model] if active_model else [])


def _risk_distribution(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total = max(len(rows), 1)
    distribution = []
    for risk in RISK_ORDER:
        risk_rows = [row for row in rows if row.get("risk_level") == risk]
        probabilities = [row.get("defect_probability") for row in risk_rows if row.get("defect_probability") is not None]
        distribution.append(
            {
                "Risk Level": risk,
                "Count": len(risk_rows),
                "Percentage": len(risk_rows) / total,
                "Avg Probability": sum(probabilities) / len(probabilities) if probabilities else 0.0,
                "Max Probability": max(probabilities) if probabilities else 0.0,
                "Min Probability": min(probabilities) if probabilities else 0.0,
            }
        )
    return distribution


def _probability_distribution(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = []
    total = max(len(rows), 1)
    probabilities = [row.get("defect_probability") or 0.0 for row in rows]
    for start in range(0, 100, 10):
        end = start + 10
        if end == 100:
            count = sum(1 for prob in probabilities if start / 100 <= prob <= 1.0)
        else:
            count = sum(1 for prob in probabilities if start / 100 <= prob < end / 100)
        buckets.append({"Bucket": f"{start}-{end}%", "Count": count, "Percentage": count / total})
    return buckets


def _top_rows(rows: list[dict[str, Any]], top_n: int) -> list[dict[str, Any]]:
    limit = max(1, min(int(top_n or 20), 100))
    return sorted(rows, key=lambda row: row.get("defect_probability") or 0.0, reverse=True)[:limit]


def _module_prediction_table(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    table = []
    for idx, row in enumerate(sorted(rows, key=lambda item: item.get("defect_probability") or 0.0, reverse=True), start=1):
        table.append(
            {
                "Rank": idx,
                "Module Name": row.get("module_name"),
                "Defect Probability": row.get("defect_probability"),
                "Prediction Label": row.get("prediction_label"),
                "Risk Level": row.get("risk_level"),
                "Risk Score": row.get("risk_score"),
                "Suggested Action": row.get("suggested_action"),
                "Model Source": row.get("model_source"),
                "Threshold": row.get("threshold"),
                "LOC": row.get("loc"),
                "NCLOC": row.get("ncloc"),
                "CLOC": row.get("cloc"),
                "Complexity": row.get("complexity"),
                "Cyclomatic Complexity": row.get("cyclomatic_complexity"),
                "Depth of Nesting": row.get("depth_of_nesting"),
                "Coupling": row.get("coupling"),
                "Cohesion": row.get("cohesion"),
                "Information Flow Complexity": row.get("information_flow_complexity"),
                "Code Churn": row.get("code_churn"),
                "Change Request Backlog": row.get("change_request_backlog"),
                "Pending Effort Hours": row.get("pending_effort_hours"),
                "Percent Reused": row.get("percent_reused"),
                "Created At": _fmt_dt(row.get("created_at")),
            }
        )
    return table


def _top_risk_table(rows: list[dict[str, Any]], top_n: int) -> list[dict[str, Any]]:
    table = []
    for idx, row in enumerate(_top_rows(rows, top_n), start=1):
        table.append(
            {
                "Rank": idx,
                "Module Name": row.get("module_name"),
                "Defect Probability": row.get("defect_probability"),
                "Risk Level": row.get("risk_level"),
                "Risk Score": row.get("risk_score"),
                "LOC": row.get("loc"),
                "Complexity": row.get("complexity"),
                "Coupling": row.get("coupling"),
                "Code Churn": row.get("code_churn"),
                "Suggested Action": row.get("suggested_action"),
            }
        )
    return table


def _heatmap_table(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    table = []
    for row in sorted(rows, key=lambda item: item.get("defect_probability") or 0.0, reverse=True):
        risk = row.get("risk_level")
        table.append(
            {
                "Module Name": row.get("module_name"),
                "Defect Probability": row.get("defect_probability"),
                "Risk Level": risk,
                "Risk Score": row.get("risk_score"),
                "LOC": row.get("loc"),
                "Complexity": row.get("complexity"),
                "Coupling": row.get("coupling"),
                "Code Churn": row.get("code_churn"),
                "Heatmap Value": row.get("defect_probability"),
                "Heatmap Color": f"#{RISK_HEAT.get(risk, '94A3B8')}",
            }
        )
    return table


def _confusion_matrix(metrics: dict[str, Any] | None) -> dict[str, int]:
    if not metrics:
        return {"tn": 0, "fp": 0, "fn": 0, "tp": 0}
    matrix = metrics.get("confusion_matrix") if isinstance(metrics, dict) else None
    if isinstance(matrix, dict):
        return {key: _safe_int(matrix.get(key)) for key in ["tn", "fp", "fn", "tp"]}
    if isinstance(matrix, list) and len(matrix) >= 2:
        return {
            "tn": _safe_int(matrix[0][0] if len(matrix[0]) > 0 else 0),
            "fp": _safe_int(matrix[0][1] if len(matrix[0]) > 1 else 0),
            "fn": _safe_int(matrix[1][0] if len(matrix[1]) > 0 else 0),
            "tp": _safe_int(matrix[1][1] if len(matrix[1]) > 1 else 0),
        }
    return {"tn": 0, "fp": 0, "fn": 0, "tp": 0}


def _model_metrics_table(models: list[dict[str, Any]]) -> list[dict[str, Any]]:
    table = []
    for model in models:
        metrics = _json_loads(model.get("metrics_json"), {}) or {}
        confusion = _confusion_matrix(metrics)
        table.append(
            {
                "Model Name": model.get("name"),
                "Active": "Yes" if model.get("is_active") else "No",
                "Best": "Yes" if model.get("is_best") else "No",
                "Model Type": model.get("model_type"),
                "Training Profile": model.get("training_profile"),
                "Threshold": model.get("threshold"),
                "Accuracy": model.get("accuracy"),
                "Precision": model.get("precision"),
                "Recall": model.get("recall"),
                "F1": model.get("f1_score"),
                "ROC-AUC": model.get("roc_auc"),
                "PR-AUC": model.get("pr_auc"),
                "TN": confusion["tn"],
                "FP": confusion["fp"],
                "FN": confusion["fn"],
                "TP": confusion["tp"],
                "Selection Strategy": model.get("selection_strategy"),
                "Selection Score": model.get("selection_score"),
                "Selection Reason": metrics.get("selection_reason") or metrics.get("best_model_selection_reason") or "",
                "Feature Columns": ", ".join(_json_loads(model.get("feature_list_json"), []) or []),
                "Created At": _fmt_dt(model.get("created_at")),
            }
        )
    return table


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def _style_header(ws, row_idx: int = 1):
    for cell in ws[row_idx]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_sheet(ws):
    thin = Side(style="thin", color=LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    _style_header(ws, 1)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 3, 11), 42)


def _write_table(ws, rows: list[dict[str, Any]], header_order: list[str] | None = None):
    if not rows:
        ws.append(["No data"])
        _style_sheet(ws)
        return
    headers = header_order or list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    _style_sheet(ws)
    _format_table(ws, headers)


def _format_table(ws, headers: list[str]):
    percent_columns = {
        "Defect Probability",
        "Defect Rate",
        "Avg Defect Probability",
        "Highest Defect Probability",
        "Lowest Defect Probability",
        "Percentage",
        "Avg Probability",
        "Max Probability",
        "Min Probability",
        "Cohesion",
        "Percent Reused",
        "Heatmap Value",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "ROC-AUC",
        "PR-AUC",
    }
    decimal_columns = {"Threshold", "Risk Score", "Selection Score"}
    risk_col = headers.index("Risk Level") + 1 if "Risk Level" in headers else None
    prediction_col = headers.index("Prediction Label") + 1 if "Prediction Label" in headers else None
    for col_idx, header in enumerate(headers, start=1):
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if header in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
            elif header in decimal_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
            elif isinstance(cell.value, float):
                cell.number_format = "0.00"
    if risk_col:
        for row_idx in range(2, ws.max_row + 1):
            risk = str(ws.cell(row_idx, risk_col).value or "").upper()
            if risk in RISK_FILLS:
                cell = ws.cell(row_idx, risk_col)
                cell.fill = PatternFill("solid", fgColor=RISK_FILLS[risk])
                cell.font = Font(bold=True, color=RISK_TEXT[risk])
    if prediction_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, prediction_col)
            cell.fill = PatternFill("solid", fgColor="FEE2E2" if int(cell.value or 0) == 1 else "DCFCE7")
            cell.font = Font(bold=True, color="991B1B" if int(cell.value or 0) == 1 else "166534")


def _write_summary_sheet(ws, report: dict[str, Any]):
    summary = report["summary"]
    rows = [
        {"Metric": "Project ID", "Value": summary["project_id"], "Value Type": "count"},
        {"Metric": "Dataset Name", "Value": summary["dataset_name"], "Value Type": "text"},
        {"Metric": "Analysis Date", "Value": _fmt_dt(summary["analysis_date"]), "Value Type": "datetime"},
        {"Metric": "Active Model", "Value": summary["active_model"], "Value Type": "text"},
        {"Metric": "Model Type", "Value": summary["model_type"], "Value Type": "text"},
        {"Metric": "Training Profile", "Value": summary["training_profile"], "Value Type": "text"},
        {"Metric": "Threshold", "Value": summary["threshold"], "Value Type": "decimal"},
        {"Metric": "Total Modules", "Value": summary["total_modules"], "Value Type": "count"},
        {"Metric": "Predicted Defect Modules", "Value": summary["predicted_defect_modules"], "Value Type": "count"},
        {"Metric": "Defects Detected", "Value": summary["predicted_defect_modules"], "Value Type": "count"},
        {"Metric": "Defect Rate", "Value": summary["defect_rate"], "Value Type": "percentage"},
        {"Metric": "Avg Defect Probability", "Value": summary["avg_defect_probability"], "Value Type": "percentage"},
        {"Metric": "Highest Defect Probability", "Value": summary["highest_defect_probability"], "Value Type": "percentage"},
        {"Metric": "Lowest Defect Probability", "Value": summary["lowest_defect_probability"], "Value Type": "percentage"},
        {"Metric": "Low Count", "Value": summary["low_count"], "Value Type": "count"},
        {"Metric": "Medium Count", "Value": summary["medium_count"], "Value Type": "count"},
        {"Metric": "High Count", "Value": summary["high_count"], "Value Type": "count"},
        {"Metric": "Critical Count", "Value": summary["critical_count"], "Value Type": "count"},
        {"Metric": "High + Critical Count", "Value": summary["high_critical_count"], "Value Type": "count"},
        {"Metric": "Model Source", "Value": summary["model_source"], "Value Type": "text"},
        {"Metric": "Report Generated At", "Value": _fmt_dt(summary["generated_at"]), "Value Type": "datetime"},
    ]
    _write_table(ws, rows, ["Metric", "Value", "Value Type"])
    for row_idx in range(2, ws.max_row + 1):
        value_type = ws.cell(row_idx, 3).value
        value = ws.cell(row_idx, 2)
        if value_type == "percentage" and isinstance(value.value, (int, float)):
            value.number_format = "0.00%"
        elif value_type == "decimal" and isinstance(value.value, (int, float)):
            value.number_format = "0.00"
        elif value_type == "count" and isinstance(value.value, (int, float)):
            value.number_format = "0"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18


def _add_bar_chart(ws, title: str, data_col: int, cat_col: int, anchor: str, max_row: int | None = None, horizontal: bool = False):
    if ws.max_row <= 1:
        return
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Count" if not horizontal else ""
    chart.x_axis.title = "" if not horizontal else "Probability"
    chart.style = 10
    chart.height = 8
    chart.width = 16
    chart.barDir = "bar" if horizontal else "col"
    max_row = max_row or ws.max_row
    data = Reference(ws, min_col=data_col, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    ws.add_chart(chart, anchor)


def _add_pie_chart(ws, title: str, data_col: int, cat_col: int, anchor: str):
    if ws.max_row <= 1:
        return
    chart = PieChart()
    chart.title = title
    chart.height = 7
    chart.width = 10
    data = Reference(ws, min_col=data_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    ws.add_chart(chart, anchor)


def dataset_xlsx(dataset_id: int, options: ExportOptions | None = None) -> bytes:
    options = options or ExportOptions()
    report = _build_report(dataset_id)
    rows = report["rows"]
    wb = Workbook()

    ws = wb.active
    ws.title = "Executive Summary"
    _write_summary_sheet(ws, report)

    module_rows = _module_prediction_table(rows)
    ws_modules = wb.create_sheet("Module Predictions")
    _write_table(ws_modules, module_rows)
    prob_col = list(module_rows[0].keys()).index("Defect Probability") + 1 if module_rows else 3
    prob_letter = get_column_letter(prob_col)
    ws_modules.conditional_formatting.add(
        f"{prob_letter}2:{prob_letter}{ws_modules.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color=RISK_HEAT["LOW"], mid_type="num", mid_value=0.6, mid_color=RISK_HEAT["MEDIUM"], end_type="num", end_value=1, end_color=RISK_HEAT["CRITICAL"]),
    )

    risk_rows = _risk_distribution(rows)
    ws_risk = wb.create_sheet("Risk Distribution")
    _write_table(ws_risk, risk_rows)
    if options.include_charts:
        _add_bar_chart(ws_risk, "Risk Distribution", data_col=2, cat_col=1, anchor="H2")
        _add_pie_chart(ws_risk, "Risk Mix", data_col=2, cat_col=1, anchor="H20")

    bucket_rows = _probability_distribution(rows)
    ws_prob = wb.create_sheet("Probability Distribution")
    _write_table(ws_prob, bucket_rows)
    if options.include_charts:
        _add_bar_chart(ws_prob, "Defect Probability Distribution", data_col=2, cat_col=1, anchor="E2")

    top_rows = _top_risk_table(rows, options.top_n)
    ws_top = wb.create_sheet("Top Risk Modules")
    _write_table(ws_top, top_rows)
    if options.include_charts and top_rows:
        _add_bar_chart(ws_top, f"Top {len(top_rows)} Risk Modules", data_col=3, cat_col=2, anchor="L2", max_row=min(ws_top.max_row, len(top_rows) + 1), horizontal=True)

    ws_heat = wb.create_sheet("Heatmap Data")
    heat_rows = _heatmap_table(rows)
    _write_table(ws_heat, heat_rows)
    if heat_rows:
        headers = list(heat_rows[0].keys())
        heat_col = headers.index("Heatmap Value") + 1
        heat_letter = get_column_letter(heat_col)
        ws_heat.conditional_formatting.add(
            f"{heat_letter}2:{heat_letter}{ws_heat.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color=RISK_HEAT["LOW"], mid_type="num", mid_value=0.6, mid_color=RISK_HEAT["MEDIUM"], end_type="num", end_value=1, end_color=RISK_HEAT["CRITICAL"]),
        )
        color_col = headers.index("Heatmap Color") + 1
        for row_idx in range(2, ws_heat.max_row + 1):
            risk = ws_heat.cell(row_idx, headers.index("Risk Level") + 1).value
            fill = RISK_HEAT.get(str(risk).upper(), "94A3B8")
            ws_heat.cell(row_idx, color_col).fill = PatternFill("solid", fgColor=fill)
            ws_heat.cell(row_idx, color_col).font = Font(color=WHITE if risk in {"HIGH", "CRITICAL"} else INK, bold=True)

    ws_model = wb.create_sheet("Model Metrics")
    model_rows = _model_metrics_table(report["models"])
    _write_table(ws_model, model_rows or [{"Model Name": "No model", "Active": "No"}])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def dataset_csv(dataset_id: int) -> bytes:
    report = _build_report(dataset_id)
    return dataframe_to_csv_bytes(pd.DataFrame(_module_prediction_table(report["rows"])))


def _paragraph_style():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=22, leading=26, textColor=colors.HexColor("#0F172A"), spaceAfter=8))
    styles.add(ParagraphStyle(name="SectionTitle", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=15, leading=18, textColor=colors.HexColor("#0F172A"), spaceBefore=10, spaceAfter=8))
    styles.add(ParagraphStyle(name="SmallMuted", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#64748B")))
    styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=7, leading=8, textColor=colors.HexColor("#0F172A")))
    styles.add(ParagraphStyle(name="CenterCell", parent=styles["BodyText"], fontSize=7, leading=8, alignment=TA_CENTER, textColor=colors.HexColor("#0F172A")))
    return styles


def _pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
    canvas.line(doc.leftMargin, 14 * mm, doc.pagesize[0] - doc.rightMargin, 14 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748B"))
    canvas.drawString(doc.leftMargin, 8 * mm, "DefectAI Analysis Report")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 8 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _pdf_table(data: list[list[Any]], header_fill: colors.Color = colors.HexColor("#0F172A"), col_widths: list[float] | None = None, repeat_rows: int = 1) -> Table:
    table = Table(data, colWidths=col_widths, repeatRows=repeat_rows, hAlign="LEFT")
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), header_fill),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEADING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
    )
    table.setStyle(style)
    return table


def _risk_chart(rows: list[dict[str, Any]]) -> Drawing:
    data = _risk_distribution(rows)
    drawing = Drawing(520, 150)
    x, baseline = 40, 25
    max_count = max((item["Count"] for item in data), default=1) or 1
    drawing.add(String(0, 132, "Risk Distribution", fontSize=12, fillColor=colors.HexColor("#0F172A"), fontName="Helvetica-Bold"))
    for item in data:
        risk = item["Risk Level"]
        height = max(4, (item["Count"] / max_count) * 90)
        drawing.add(Rect(x, baseline, 62, height, fillColor=colors.HexColor(f"#{RISK_HEAT[risk]}"), strokeColor=colors.HexColor("#CBD5E1")))
        drawing.add(String(x + 8, baseline + height + 6, str(item["Count"]), fontSize=8, fillColor=colors.HexColor("#0F172A")))
        drawing.add(String(x + 2, 8, risk, fontSize=7, fillColor=PDF_RISK_TEXT[risk], fontName="Helvetica-Bold"))
        x += 92
    return drawing


def _probability_chart(rows: list[dict[str, Any]]) -> Drawing:
    buckets = _probability_distribution(rows)
    drawing = Drawing(520, 150)
    max_count = max((item["Count"] for item in buckets), default=1) or 1
    drawing.add(String(0, 132, "Probability Distribution", fontSize=12, fillColor=colors.HexColor("#0F172A"), fontName="Helvetica-Bold"))
    x, baseline = 18, 25
    for item in buckets:
        height = max(3, (item["Count"] / max_count) * 88)
        drawing.add(Rect(x, baseline, 32, height, fillColor=colors.HexColor("#60A5FA"), strokeColor=colors.HexColor("#CBD5E1")))
        drawing.add(String(x + 4, baseline + height + 5, str(item["Count"]), fontSize=6, fillColor=colors.HexColor("#0F172A")))
        drawing.add(String(x - 2, 8, item["Bucket"].replace("-", "-\n")[:6], fontSize=5, fillColor=colors.HexColor("#475569")))
        x += 48
    return drawing


def _top_risk_chart(rows: list[dict[str, Any]], top_n: int) -> Drawing:
    top = _top_rows(rows, min(top_n, 10))
    drawing = Drawing(520, 190)
    drawing.add(String(0, 172, f"Top {len(top)} Risk Modules", fontSize=12, fillColor=colors.HexColor("#0F172A"), fontName="Helvetica-Bold"))
    y = 150
    for row in top:
        prob = row.get("defect_probability") or 0.0
        width = max(6, prob * 310)
        risk = row.get("risk_level")
        drawing.add(String(0, y + 2, str(row.get("module_name", ""))[:32], fontSize=7, fillColor=colors.HexColor("#334155")))
        drawing.add(Rect(160, y, width, 10, fillColor=colors.HexColor(f"#{RISK_HEAT.get(risk, '60A5FA')}"), strokeColor=None))
        drawing.add(String(160 + width + 6, y + 1, f"{prob * 100:.1f}%", fontSize=7, fillColor=colors.HexColor("#0F172A")))
        y -= 15
    return drawing


def _heatmap_pdf_table(rows: list[dict[str, Any]], top_n: int, styles) -> Table:
    cells = []
    for row in _top_rows(rows, min(max(top_n, 20), 50)):
        risk = row.get("risk_level")
        text = f"<b>{_safe_text(row.get('module_name'), '')[:22]}</b><br/>{_fmt_pct(row.get('defect_probability'))} | {risk}"
        cells.append((text, risk))
    columns = 5
    data = []
    for idx in range(0, len(cells), columns):
        chunk = cells[idx : idx + columns]
        data.append([Paragraph(text, styles["CenterCell"]) for text, _ in chunk] + [""] * (columns - len(chunk)))
    if not data:
        data = [["No heatmap data"]]
    table = Table(data, colWidths=[95] * columns, hAlign="LEFT")
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]
    for row_idx, row in enumerate(data):
        for col_idx, _ in enumerate(row):
            source_idx = row_idx * columns + col_idx
            if source_idx < len(cells):
                risk = cells[source_idx][1]
                commands.append(("BACKGROUND", (col_idx, row_idx), (col_idx, row_idx), PDF_RISK_FILL.get(risk, colors.HexColor("#F1F5F9"))))
    table.setStyle(TableStyle(commands))
    return table


def _pdf_summary_table(summary: dict[str, Any]) -> Table:
    data = [
        ["Metric", "Value", "Metric", "Value"],
        ["Total Modules", summary["total_modules"], "Predicted Defect Modules", summary["predicted_defect_modules"]],
        ["Defect Rate", _fmt_pct(summary["defect_rate"]), "Avg Defect Probability", _fmt_pct(summary["avg_defect_probability"])],
        ["High Count", summary["high_count"], "Critical Count", summary["critical_count"]],
        ["High + Critical Count", summary["high_critical_count"], "Threshold", _fmt_number(summary["threshold"], 2)],
        ["Active Model", summary["active_model"], "Model Type", summary["model_type"]],
    ]
    table = _pdf_table(data, col_widths=[95, 160, 120, 160])
    table.setStyle(TableStyle([("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFFFF"))]))
    return table


def _pdf_model_performance(model: dict[str, Any] | None) -> list[list[Any]]:
    if not model:
        return [["Metric", "Value"], ["Active Model", "N/A"]]
    metrics = _json_loads(model.get("metrics_json"), {}) or {}
    confusion = _confusion_matrix(metrics)
    return [
        ["Metric", "Value", "Metric", "Value"],
        ["Accuracy", _fmt_pct(model.get("accuracy")), "Precision", _fmt_pct(model.get("precision"))],
        ["Recall", _fmt_pct(model.get("recall")), "F1", _fmt_pct(model.get("f1_score"))],
        ["ROC-AUC", _fmt_number(model.get("roc_auc"), 3), "PR-AUC", _fmt_number(model.get("pr_auc"), 3)],
        ["Threshold", _fmt_number(model.get("threshold"), 2), "Training Profile", _safe_text(model.get("training_profile"))],
        ["TN", confusion["tn"], "FP", confusion["fp"]],
        ["FN", confusion["fn"], "TP", confusion["tp"]],
    ]


def dataset_pdf(dataset_id: int, options: ExportOptions | None = None) -> bytes:
    options = options or ExportOptions()
    report = _build_report(dataset_id)
    rows = report["rows"]
    summary = report["summary"]
    styles = _paragraph_style()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title=f"DefectAI Analysis Report - {summary['dataset_name']}",
    )
    story = [
        Paragraph("DefectAI Analysis Report", styles["ReportTitle"]),
        Paragraph(
            f"Dataset: <b>{summary['dataset_name']}</b> &nbsp;&nbsp; Project #{summary['project_id']} &nbsp;&nbsp; Analysis: {_fmt_dt(summary['analysis_date'])} &nbsp;&nbsp; Generated: {_fmt_dt(summary['generated_at'])}",
            styles["SmallMuted"],
        ),
        Spacer(1, 8),
        Paragraph("Executive Summary", styles["SectionTitle"]),
        _pdf_summary_table(summary),
        Spacer(1, 12),
    ]

    if options.include_charts:
        story.extend(
            [
                Paragraph("Risk Distribution", styles["SectionTitle"]),
                _risk_chart(rows),
                Spacer(1, 8),
                Paragraph("Defect Probability Overview", styles["SectionTitle"]),
                _probability_chart(rows),
                Spacer(1, 8),
                _top_risk_chart(rows, options.top_n),
            ]
        )

    top_table_data = [["Rank", "Module", "Probability", "Risk", "Risk Score", "Action"]]
    for idx, row in enumerate(_top_rows(rows, options.top_n), start=1):
        top_table_data.append(
            [
                idx,
                Paragraph(_safe_text(row.get("module_name"), "")[:44], styles["Cell"]),
                _fmt_pct(row.get("defect_probability")),
                row.get("risk_level"),
                _fmt_pct(row.get("risk_score")),
                Paragraph(_safe_text(row.get("suggested_action"), "")[:85], styles["Cell"]),
            ]
        )
    story.extend([Spacer(1, 10), Paragraph("Top Risk Modules", styles["SectionTitle"]), _pdf_table(top_table_data, col_widths=[32, 210, 70, 70, 70, 245])])

    if options.include_heatmap:
        legend = [["LOW", "MEDIUM", "HIGH", "CRITICAL"]]
        legend_table = Table(legend, colWidths=[70, 70, 70, 70], hAlign="LEFT")
        legend_style = [("GRID", (0, 0), (-1, -1), 0.4, colors.white), ("FONTSIZE", (0, 0), (-1, -1), 7), ("ALIGN", (0, 0), (-1, -1), "CENTER")]
        for idx, risk in enumerate(RISK_ORDER):
            legend_style.append(("BACKGROUND", (idx, 0), (idx, 0), PDF_RISK_FILL[risk]))
            legend_style.append(("TEXTCOLOR", (idx, 0), (idx, 0), PDF_RISK_TEXT[risk]))
        legend_table.setStyle(TableStyle(legend_style))
        story.extend([PageBreak(), Paragraph("Risk Heatmap", styles["SectionTitle"]), legend_table, Spacer(1, 8), _heatmap_pdf_table(rows, options.top_n, styles)])

    story.extend(
        [
            PageBreak(),
            Paragraph("Model Performance", styles["SectionTitle"]),
            _pdf_table(_pdf_model_performance(report["model"]), col_widths=[100, 160, 110, 160]),
        ]
    )

    if options.include_full_modules:
        story.extend([PageBreak(), Paragraph("Full Module Detail Appendix", styles["SectionTitle"])])
        appendix = [["Rank", "Module", "Probability", "Label", "Risk", "Risk Score", "Threshold", "LOC", "Complexity", "Coupling", "Churn", "Action"]]
        for idx, row in enumerate(sorted(rows, key=lambda item: item.get("defect_probability") or 0.0, reverse=True), start=1):
            appendix.append(
                [
                    idx,
                    Paragraph(_safe_text(row.get("module_name"), "")[:38], styles["Cell"]),
                    _fmt_pct(row.get("defect_probability")),
                    row.get("prediction_label"),
                    row.get("risk_level"),
                    _fmt_pct(row.get("risk_score")),
                    _fmt_number(row.get("threshold"), 2),
                    _safe_int(row.get("loc")),
                    _fmt_number(row.get("complexity"), 1),
                    _fmt_number(row.get("coupling"), 1),
                    _fmt_number(row.get("code_churn"), 1),
                    Paragraph(_safe_text(row.get("suggested_action"), "")[:70], styles["Cell"]),
                ]
            )
        story.append(_pdf_table(appendix, col_widths=[28, 155, 60, 35, 50, 55, 50, 45, 55, 55, 45, 155]))

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    return buffer.getvalue()


def report_rows(report_id: int) -> list[dict[str, Any]]:
    report = report_repository.get_report(report_id)
    if not report:
        raise ValueError("Report not found")
    filters = json.loads(report.get("filters_json") or "{}")
    dataset_id = filters.get("dataset_id")
    if dataset_id:
        return _module_prediction_table(_build_report(int(dataset_id))["rows"])
    raise ValueError("Report is not linked to a dataset")


def report_csv(report_id: int) -> bytes:
    report = report_repository.get_report(report_id)
    if not report:
        return dataset_csv(report_id)
    filters = json.loads(report.get("filters_json") or "{}") if report else {}
    dataset_id = filters.get("dataset_id")
    return dataset_csv(int(dataset_id)) if dataset_id else dataframe_to_csv_bytes(pd.DataFrame(report_rows(report_id)))


def report_xlsx(report_id: int, options: ExportOptions | None = None) -> bytes:
    report = report_repository.get_report(report_id)
    if not report:
        return dataset_xlsx(report_id, options)
    filters = json.loads(report.get("filters_json") or "{}") if report else {}
    dataset_id = filters.get("dataset_id")
    if not dataset_id:
        raise ValueError("Report is not linked to a dataset")
    return dataset_xlsx(int(dataset_id), options)


def report_pdf(report_id: int, options: ExportOptions | None = None) -> bytes:
    report = report_repository.get_report(report_id)
    if not report:
        return dataset_pdf(report_id, options)
    filters = json.loads(report.get("filters_json") or "{}")
    dataset_id = filters.get("dataset_id")
    if not dataset_id:
        raise ValueError("Report is not linked to a dataset")
    return dataset_pdf(int(dataset_id), options)


def export_multiple_reports(report_ids: list[int], format: str = "csv") -> bytes:
    import zipfile

    if len(report_ids) == 1:
        if format == "csv":
            return report_csv(report_ids[0])
        if format == "xlsx":
            return report_xlsx(report_ids[0])
        if format == "pdf":
            return report_pdf(report_ids[0])
        return report_csv(report_ids[0])

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for report_id in report_ids:
            try:
                if format == "csv":
                    content = report_csv(report_id)
                    filename = f"report_{report_id}.csv"
                elif format == "xlsx":
                    content = report_xlsx(report_id)
                    filename = f"report_{report_id}.xlsx"
                elif format == "pdf":
                    content = report_pdf(report_id)
                    filename = f"report_{report_id}.pdf"
                else:
                    content = report_csv(report_id)
                    filename = f"report_{report_id}.csv"
                zip_file.writestr(filename, content)
            except Exception:
                continue
    return buffer.getvalue()
